#!/usr/bin/env python3
"""
광고비 정산 체크 도구.

사용:
    python3 ~/tools/settlement_check.py                     # 최신 폴더 자동 선택
    python3 ~/tools/settlement_check.py "<폴더경로>"        # 특정 폴더 지정
    python3 ~/tools/settlement_check.py --no-pdf            # PDF 변환 건너뛰기
    python3 ~/tools/settlement_check.py --tolerance 0.1     # RAW 비교 허용 오차 (기본 0.1 = 10%)

흐름:
    1) 검수확인서.xlsx 파싱 → 예산별 금액 표 + 총합
    2) 정산 RAW.xlsx 비교 → 매체 시트 비용 합산해서 ±오차 검증
    3) 검수확인서 + 발주서 → PDF 변환 (AppleScript / Word·Excel)
    4) _체크시트.html 생성 → 자동으로 브라우저에서 열림
"""
from __future__ import annotations

import argparse
import html
import os
import re
import subprocess
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl

SETTLEMENT_ROOT = Path.home() / "Documents" / "정산 자료"


@dataclass
class BudgetLine:
    no: int
    name: str          # 예산명
    code: str          # 예산코드
    unit_price: int    # 단가
    amount: int        # 검수금액


@dataclass
class Parsed:
    folder: Path
    inspection_xlsx: Path
    raw_xlsx: Path
    purchase_docx: Path
    period_label: str
    budget_lines: list[BudgetLine]
    grand_total: int
    raw_total: int
    raw_breakdown: dict[str, int]
    raw_warnings: list[str]


# ── 파일 검색 ──────────────────────────────────────────────

def latest_folder() -> Path:
    if not SETTLEMENT_ROOT.exists():
        sys.exit(f"❌ 폴더 없음: {SETTLEMENT_ROOT}")
    candidates = [p for p in SETTLEMENT_ROOT.iterdir() if p.is_dir() and not p.name.startswith(".")]
    if not candidates:
        sys.exit(f"❌ 정산 폴더가 없음: {SETTLEMENT_ROOT}")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s)


def find_files(folder: Path) -> tuple[Path, Path, Path]:
    # 폴더 + 하위 한 단계까지 (첨부 같은 서브폴더 케이스 대응)
    files: list[Path] = []
    for p in folder.iterdir():
        if p.name.startswith("~$") or p.name.startswith("."):
            continue
        if p.is_file():
            files.append(p)
        elif p.is_dir():
            for sub in p.iterdir():
                if sub.is_file() and not sub.name.startswith("~$") and not sub.name.startswith("."):
                    files.append(sub)

    def pick(pattern: str, kind: str) -> Path:
        rx = re.compile(pattern)
        hits = [f for f in files if rx.search(_nfc(f.name))]
        if not hits:
            sys.exit(f"❌ {kind} 파일을 찾지 못함 (패턴: {pattern})\n   폴더: {folder}")
        # 부모 폴더가 본 폴더인 걸 우선
        hits.sort(key=lambda f: 0 if f.parent == folder else 1)
        return hits[0]

    inspection = pick(r"검수\s*확인서.*\.xlsx$", "검수확인서")
    raw = pick(r"정산\s*RAW.*\.xlsx$", "정산 RAW")
    purchase = pick(r"발주서.*\.docx$", "발주서")
    return inspection, raw, purchase


# ── 검수확인서 파싱 ──────────────────────────────────────────

def parse_inspection(path: Path) -> tuple[list[BudgetLine], int, str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    period_label = wb.sheetnames[0]

    lines: list[BudgetLine] = []
    grand_total = 0

    for row in ws.iter_rows(values_only=True):
        # 헤더 행: (None, 'No', '대상월', '업무내역', '예산명', '예산코드', '단 가', '수 량', '검 수 금 액', ...)
        if row[1] == "No":
            continue
        if isinstance(row[1], int) and isinstance(row[8], (int, float)):
            lines.append(
                BudgetLine(
                    no=row[1],
                    name=str(row[4]).strip() if row[4] else "",
                    code=str(row[5]).strip() if row[5] else "",
                    unit_price=int(row[6] or 0),
                    amount=int(row[8] or 0),
                )
            )
        elif row[1] == "총 합계" and isinstance(row[8], (int, float)):
            grand_total = int(row[8])

    if not grand_total and lines:
        grand_total = sum(b.amount for b in lines)

    return lines, grand_total, period_label


# ── RAW 비교 ──────────────────────────────────────────────
# 계산시트 탭에 매체×구분 피벗이 있음.
# 컬럼 F-G 영역에 '행 레이블 / 합계 : 비용(vat,마크업 포함)' 표가 있고,
# 마지막 행이 '총합계'. 그걸 기준 총액으로 사용.

def parse_raw(path: Path) -> tuple[int, dict[str, int], list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    warnings: list[str] = []

    if "계산시트" not in wb.sheetnames:
        warnings.append("계산시트 탭 없음 — RAW 비교 불가")
        return 0, {}, warnings

    ws = wb["계산시트"]
    breakdown: dict[str, int] = {}
    grand_total = 0
    pivot_label_col: int | None = None  # '행 레이블' 컬럼 인덱스
    pivot_value_col: int | None = None  # 그 다음 숫자 컬럼

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        # 아직 피벗 못 찾았으면 '행 레이블' 마커 탐색
        if pivot_label_col is None:
            for i, c in enumerate(row):
                if isinstance(c, str) and "행 레이블" in c:
                    pivot_label_col = i
                    pivot_value_col = i + 1
                    break
            continue

        # 피벗 내부: 고정된 두 컬럼만 읽음
        cat_cell = row[pivot_label_col] if len(row) > pivot_label_col else None
        val_cell = row[pivot_value_col] if len(row) > pivot_value_col else None

        if isinstance(val_cell, str) and "#REF" in val_cell:
            warnings.append("계산시트 피벗에 #REF! 발견")
            continue

        cat = str(cat_cell).strip() if isinstance(cat_cell, str) and cat_cell.strip() else None
        val = int(val_cell) if isinstance(val_cell, (int, float)) else None

        if cat is None and val is None:
            # 빈 행이면 피벗 끝
            if breakdown:
                break
            continue

        if cat == "총합계":
            grand_total = val or 0
            break
        if cat and val is not None:
            breakdown[cat] = val

    if not grand_total and breakdown:
        grand_total = sum(breakdown.values())

    if not grand_total:
        warnings.append("계산시트에서 총합계 행을 찾지 못함")

    return grand_total, breakdown, warnings


# ── PDF 변환 (AppleScript) ───────────────────────────────────

WORD_TO_PDF_AS = '''
tell application "Microsoft Word"
    set wasRunning to running
    activate
    set theDoc to open file name "{src}" with read only
    save as theDoc file name "{dst}" file format format PDF
    close theDoc saving no
    if not wasRunning then quit saving no
end tell
'''

EXCEL_TO_PDF_AS = '''
tell application "Microsoft Excel"
    set wasRunning to running
    activate
    open "{src}"
    delay 1
    set theBook to active workbook
    set theSheet to active sheet of theBook
    save theSheet in "{dst}" as PDF file format
    close theBook saving no
    if not wasRunning then quit saving no
end tell
'''


def run_osascript(script: str) -> tuple[bool, str]:
    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=120,
        )
        return result.returncode == 0, (result.stderr or result.stdout).strip()
    except subprocess.TimeoutExpired:
        return False, "변환 시간 초과 (120초)"


def to_pdf(src: Path) -> tuple[Path | None, str]:
    dst = src.with_suffix(".pdf")
    if dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        return dst, "이미 존재 (스킵)"
    suffix = src.suffix.lower()
    if suffix == ".docx":
        script = WORD_TO_PDF_AS.format(src=src, dst=dst)
    elif suffix == ".xlsx":
        script = EXCEL_TO_PDF_AS.format(src=src, dst=dst)
    else:
        return None, f"지원하지 않는 형식: {suffix}"
    ok, msg = run_osascript(script)
    if ok and dst.exists():
        return dst, "변환 완료"
    return None, msg or "알 수 없는 오류"


# ── 체크시트 HTML 생성 ─────────────────────────────────────

CHECKSHEET_HTML = """<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>정산 체크시트 — {period}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro", "Pretendard", sans-serif;
    background: #f5f5f7; color: #1d1d1f; padding: 40px 24px; line-height: 1.5;
  }}
  .wrap {{ max-width: 980px; margin: 0 auto; }}
  h1 {{ font-size: 28px; font-weight: 700; letter-spacing: -0.02em; margin-bottom: 4px; }}
  .sub {{ color: #6e6e73; font-size: 14px; margin-bottom: 32px; }}
  .panel {{
    background: white; border-radius: 16px; padding: 24px;
    box-shadow: 0 1px 2px rgba(0,0,0,.04), 0 4px 12px rgba(0,0,0,.04);
    margin-bottom: 20px;
  }}
  .panel h2 {{ font-size: 13px; font-weight: 600; text-transform: uppercase;
    color: #6e6e73; letter-spacing: 0.06em; margin-bottom: 16px; }}
  .row {{ display: flex; justify-content: space-between; padding: 10px 0;
    border-bottom: 1px solid #f0f0f3; font-size: 14px; }}
  .row:last-child {{ border-bottom: none; }}
  .row .k {{ color: #6e6e73; }}
  .row .v {{ font-variant-numeric: tabular-nums; font-weight: 500; }}
  .status {{ display: inline-flex; align-items: center; gap: 8px;
    padding: 6px 12px; border-radius: 999px; font-size: 13px; font-weight: 600; }}
  .ok {{ background: #e8f5ee; color: #0a7a3b; }}
  .warn {{ background: #fef4e6; color: #a55a00; }}
  .err {{ background: #fde9ea; color: #b3261e; }}
  .check-callout {{
    background: #fef4e6; border-left: 4px solid #f5a623;
    padding: 14px 18px; border-radius: 10px; font-size: 14px; color: #6b4400;
    margin-bottom: 20px;
  }}
  .cards {{ display: grid; gap: 14px; }}
  .card {{ background: white; border-radius: 14px; padding: 20px;
    border: 1px solid #e8e8ed; transition: opacity .25s, border-color .25s; position: relative; }}
  .card[data-state="done"] {{ opacity: 0.55; border-color: #34c759; }}
  .card[data-state="done"]::before {{
    content: "✓ 완료"; position: absolute; top: 14px; right: 16px;
    background: #34c759; color: white; padding: 4px 10px; border-radius: 999px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.05em;
  }}
  .card .num {{ font-size: 11px; font-weight: 700; color: #ff6b35;
    letter-spacing: 0.1em; margin-bottom: 6px; }}
  .card[data-state="done"] .num {{ color: #34c759; }}
  .card .name {{ font-size: 16px; font-weight: 600; margin-bottom: 4px; }}
  .card .code {{ font-size: 12px; color: #6e6e73; font-family: SF Mono, monospace; margin-bottom: 16px; }}
  .field {{ display: flex; align-items: center; justify-content: space-between;
    background: #f5f5f7; border-radius: 10px; padding: 10px 14px; margin-bottom: 8px; }}
  .field .lbl {{ font-size: 13px; color: #6e6e73; }}
  .field .val {{ font-size: 18px; font-weight: 600; font-variant-numeric: tabular-nums; }}
  .field-quantity {{ font-size: 13px; color: #6e6e73; padding: 6px 14px; }}
  .field-quantity b {{ color: #1d1d1f; font-weight: 600; }}
  button.copy {{
    background: #007aff; color: white; border: none; padding: 6px 14px;
    border-radius: 8px; font-size: 13px; font-weight: 500; cursor: pointer;
    transition: background .15s; min-width: 96px;
  }}
  button.copy:hover {{ background: #005bb5; }}
  button.copy[data-state="copied"] {{ background: #ff9500; }}
  button.copy[data-state="copied"]:hover {{ background: #cc7700; }}
  button.copy[data-state="done"] {{ background: #34c759; }}
  .progress {{ font-size: 13px; color: #6e6e73; margin-top: -8px; margin-bottom: 16px; }}
  .progress b {{ color: #1d1d1f; font-weight: 600; }}
  .attach {{ font-size: 12px; color: #6e6e73; margin-top: 12px; }}
  .attach b {{ color: #1d1d1f; font-weight: 600; }}
  .breakdown {{ font-size: 12px; color: #6e6e73; margin-top: 8px; }}
  .breakdown table {{ margin-top: 4px; border-collapse: collapse; }}
  .breakdown td {{ padding: 2px 12px 2px 0; }}
  .breakdown td.r {{ text-align: right; font-variant-numeric: tabular-nums; }}
</style>
</head>
<body>
<div class="wrap">
  <h1>광고비 정산 체크시트</h1>
  <div class="sub">{period} · {folder_name}</div>

  <div class="panel">
    <h2>1. 검증 (RAW vs 검수확인서)</h2>
    <div class="row">
      <span class="k">검수확인서 총합 <small style="color:#999">(부가세 별도)</small></span>
      <span class="v">{grand_total:,} 원</span>
    </div>
    <div class="row">
      <span class="k">× 1.10 부가세 적용 <small style="color:#999">(예상 RAW)</small></span>
      <span class="v">{expected_raw:,} 원</span>
    </div>
    <div class="row">
      <span class="k">실제 RAW 합계 <small style="color:#999">(계산시트 vat,마크업 포함)</small></span>
      <span class="v">{raw_total:,} 원</span>
    </div>
    <div class="row">
      <span class="k">차이 <small style="color:#999">(RAW − 예상)</small></span>
      <span class="v">{diff:+,} 원 ({diff_pct:+.2f}%)</span>
    </div>
    <div class="row" style="margin-top:8px;">
      <span class="k">상태</span>
      <span class="status {status_class}">{status_text}</span>
    </div>
    {warnings_html}
    <div class="breakdown">
      <div>RAW 구분별 (계산시트 피벗):</div>
      <table>{breakdown_rows}</table>
    </div>
  </div>

  <div class="check-callout">
    <b>👁️ 발주서 PDF 하단 견적서 이미지</b> 금액이 <b>{grand_total:,} 원</b> 인지 1초 눈 확인.
  </div>

  <div class="panel">
    <h2>2. 구매시스템 입력 — 예산별 카드</h2>
    <div class="sub" style="margin-bottom:8px;">검수단가에 카드 금액 입력, 검수수량은 모두 <b>1</b>, 첨부파일 3개 업로드 후 승인.</div>
    <div class="progress">진행: <b id="progress-count">0</b> / {n_cards} 완료</div>
    <div class="sub" style="margin-bottom:16px; font-size:12px; color:#999;">금액 버튼 1번 클릭 = 복사 · 2번 클릭 = 완료 표시 · 3번 클릭 = 되돌리기</div>
    <div class="cards">
      {cards_html}
    </div>
  </div>

  <div class="panel">
    <h2>3. 첨부 파일 (모든 카드 공통)</h2>
    <div class="row"><span class="k">검수확인서 PDF</span><span class="v" style="font-size:12px;">{insp_pdf}</span></div>
    <div class="row"><span class="k">발주서 PDF</span><span class="v" style="font-size:12px;">{po_pdf}</span></div>
    <div class="row"><span class="k">정산 RAW xlsx</span><span class="v" style="font-size:12px;">{raw_xlsx}</span></div>
  </div>
</div>

<script>
const STORAGE_KEY = "settlement-checksheet:{storage_key}";

const BUTTON_LABEL = {{
  idle: "복사",
  copied: "복사됨 ✓",
  done: "완료 ✓"
}};

function loadState() {{
  try {{ return JSON.parse(localStorage.getItem(STORAGE_KEY) || "{{}}"); }}
  catch (e) {{ return {{}}; }}
}}

function saveState(state) {{
  localStorage.setItem(STORAGE_KEY, JSON.stringify(state));
}}

function applyCardState(card, cardState) {{
  card.dataset.state = cardState === "done" ? "done" : "idle";
  const btn = card.querySelector("button.copy");
  if (!btn) return;
  btn.dataset.state = cardState;
  btn.textContent = BUTTON_LABEL[cardState] || BUTTON_LABEL.idle;
}}

function updateProgress() {{
  const done = document.querySelectorAll('.card[data-state="done"]').length;
  document.getElementById("progress-count").textContent = done;
}}

function onCopyClick(btn) {{
  const card = btn.closest(".card");
  const cardId = card.dataset.cardId;
  const state = loadState();
  const current = btn.dataset.state || "idle";

  // 사이클: idle → copied → done → idle
  let next;
  if (current === "idle") {{
    next = "copied";
    navigator.clipboard.writeText(btn.dataset.amount);
  }} else if (current === "copied") {{
    next = "done";
  }} else {{
    next = "idle";
  }}

  applyCardState(card, next);
  state[cardId] = next;
  saveState(state);
  updateProgress();
}}

// 페이지 로드 시 저장된 상태 복원
document.addEventListener("DOMContentLoaded", () => {{
  const state = loadState();
  document.querySelectorAll(".card").forEach(card => {{
    const saved = state[card.dataset.cardId] || "idle";
    applyCardState(card, saved);
  }});
  updateProgress();
}});
</script>
</body>
</html>
"""


def make_card_html(line: BudgetLine, total: int) -> str:
    return f"""
      <div class="card" data-card-id="{html.escape(line.code)}" data-state="idle">
        <div class="num">예산 {line.no} / {total}</div>
        <div class="name">{html.escape(line.name)}</div>
        <div class="code">{html.escape(line.code)}</div>
        <div class="field">
          <span class="lbl">검수단가</span>
          <span class="val">{line.amount:,} 원</span>
          <button class="copy" data-state="idle" data-amount="{line.amount}" onclick="onCopyClick(this)">복사</button>
        </div>
        <div class="field-quantity">검수수량: <b>1</b> (고정)</div>
      </div>
    """


def make_checksheet(parsed: Parsed, pdf_paths: dict[str, Path | None], tolerance: float) -> Path:
    # RAW은 VAT+마크업 포함, 검수확인서는 VAT 별도 → 보정 후 비교
    expected_raw = int(round(parsed.grand_total * 1.10))
    diff = parsed.raw_total - expected_raw
    diff_pct = (diff / expected_raw * 100) if expected_raw else 0.0

    if abs(diff_pct) <= tolerance * 100 and not parsed.raw_warnings:
        status_class, status_text = "ok", f"✅ 정상 (VAT 보정 후 ±{tolerance*100:.0f}% 이내)"
    elif abs(diff_pct) <= tolerance * 200:
        status_class, status_text = "warn", "⚠️ 차이 있음 — 한번 확인"
    else:
        status_class, status_text = "err", "❌ 차이 큼 — 점검 필요"

    warnings_html = ""
    if parsed.raw_warnings:
        warns = "".join(f"<li>{html.escape(w)}</li>" for w in parsed.raw_warnings)
        warnings_html = f'<div class="row"><span class="k">경고</span><span class="v"><ul style="margin:0;padding-left:18px;color:#a55a00;">{warns}</ul></span></div>'

    breakdown_rows = "".join(
        f"<tr><td>{html.escape(k)}</td><td class='r'>{v:,}</td></tr>"
        for k, v in sorted(parsed.raw_breakdown.items(), key=lambda x: -x[1])
        if v > 0
    )

    cards_html = "\n".join(make_card_html(b, len(parsed.budget_lines)) for b in parsed.budget_lines)

    insp_pdf = str(pdf_paths.get("inspection") or "❌ 변환 실패")
    po_pdf = str(pdf_paths.get("purchase") or "❌ 변환 실패")

    html_body = CHECKSHEET_HTML.format(
        period=html.escape(parsed.period_label),
        folder_name=html.escape(parsed.folder.name),
        storage_key=html.escape(parsed.folder.name),
        n_cards=len(parsed.budget_lines),
        grand_total=parsed.grand_total,
        expected_raw=expected_raw,
        raw_total=parsed.raw_total,
        diff=diff,
        diff_pct=diff_pct,
        status_class=status_class,
        status_text=status_text,
        warnings_html=warnings_html,
        breakdown_rows=breakdown_rows,
        cards_html=cards_html,
        insp_pdf=html.escape(insp_pdf),
        po_pdf=html.escape(po_pdf),
        raw_xlsx=html.escape(str(parsed.raw_xlsx)),
    )

    out = parsed.folder / "_체크시트.html"
    out.write_text(html_body, encoding="utf-8")
    return out


# ── 메인 ─────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="광고비 정산 체크 도구")
    ap.add_argument("folder", nargs="?", help="정산 폴더 경로 (생략 시 최신 자동 선택)")
    ap.add_argument("--no-pdf", action="store_true", help="PDF 변환 건너뛰기")
    ap.add_argument("--no-open", action="store_true", help="브라우저 자동 오픈 끄기")
    ap.add_argument("--tolerance", type=float, default=0.02, help="VAT 보정 후 RAW 비교 허용 오차 (기본 0.02 = 2%)")
    args = ap.parse_args()

    folder = Path(args.folder).expanduser().resolve() if args.folder else latest_folder()
    print(f"📂 폴더: {folder.name}")

    inspection, raw, purchase = find_files(folder)
    print(f"  • 검수확인서: {inspection.name}")
    print(f"  • 정산 RAW:   {raw.name}")
    print(f"  • 발주서:     {purchase.name}")

    print("\n📄 검수확인서 파싱…")
    budget_lines, grand_total, period = parse_inspection(inspection)
    print(f"  예산 {len(budget_lines)}건, 총합 {grand_total:,} 원")

    print("\n🔎 RAW 비교…")
    raw_total, breakdown, warnings = parse_raw(raw)
    expected = int(round(grand_total * 1.10))
    diff = raw_total - expected
    diff_pct = (diff / expected * 100) if expected else 0
    print(f"  검수확인서 ×1.1 = {expected:,} 원  /  RAW 합계 = {raw_total:,} 원")
    print(f"  차이 {diff:+,} ({diff_pct:+.2f}%)")
    if warnings:
        for w in warnings:
            print(f"  ⚠️  {w}")

    parsed = Parsed(
        folder=folder,
        inspection_xlsx=inspection,
        raw_xlsx=raw,
        purchase_docx=purchase,
        period_label=period,
        budget_lines=budget_lines,
        grand_total=grand_total,
        raw_total=raw_total,
        raw_breakdown=breakdown,
        raw_warnings=warnings,
    )

    pdf_paths: dict[str, Path | None] = {"inspection": None, "purchase": None}
    if not args.no_pdf:
        print("\n📑 PDF 변환…")
        for key, src in (("inspection", inspection), ("purchase", purchase)):
            print(f"  • {src.name} …", end=" ", flush=True)
            pdf, msg = to_pdf(src)
            pdf_paths[key] = pdf
            print(f"{'✓' if pdf else '✗'} {msg}")

    print("\n📋 체크시트 생성…")
    out = make_checksheet(parsed, pdf_paths, args.tolerance)
    print(f"  → {out}")

    if not args.no_open:
        subprocess.run(["open", str(out)])
        print("\n🌐 브라우저에서 열림.")


if __name__ == "__main__":
    main()
